from __future__ import annotations

import argparse
import json
import os
import sqlite3
import datetime
import logging
from typing import List, Dict

from openpyxl import Workbook, load_workbook

from config import DB_CONFIG, RESULT_DIR
from db_client import DBClient
from query_engine import QueryEngine

logger = logging.getLogger(__name__)


def _parse_question_cell(value: str) -> List[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    try:
        arr = json.loads(text)
        return [str(x.get("Q", "")).strip() for x in arr if str(x.get("Q", "")).strip()]
    except Exception:
        return [text]


def _iter_question_rows(input_xlsx: str):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in range(2, ws.max_row + 1):
        qid = ws.cell(r, 1).value
        qtype = ws.cell(r, 2).value
        qraw = ws.cell(r, 3).value
        if not qid:
            continue
        yield str(qid), str(qtype or ""), qraw


def run_batch(input_files: List[str], output_xlsx: str, result_dir: str):

    db = DBClient(DB_CONFIG)
    engine = QueryEngine(db, result_dir)

    out_rows: List[Dict] = []
    os.makedirs(result_dir, exist_ok=True)

    # prepare failure persistence
    failed_db_path = os.path.join(result_dir, "failed_jobs.sqlite")
    conn_fail = sqlite3.connect(failed_db_path)
    cur_fail = conn_fail.cursor()
    cur_fail.execute(
        """
        CREATE TABLE IF NOT EXISTS failed_jobs(
            question_id TEXT PRIMARY KEY,
            question_type TEXT,
            question TEXT,
            status TEXT,
            error_type TEXT,
            error_message TEXT,
            sql TEXT,
            attempts INTEGER,
            last_attempt_ts TEXT
        )
        """
    )
    conn_fail.commit()

    try:
        for input_xlsx in input_files:
            for qid, qtype, qraw in _iter_question_rows(input_xlsx):
                q_list = _parse_question_cell(qraw)
                q_turns = q_list if q_list else [str(qraw or "")]
                context = {}
                turn_answers = []
                attempts = 0
                last_error = ""
                last_sql = ""
                failed_flag = False

                for idx, q_text in enumerate(q_turns, start=1):
                    attempts += 1
                    try:
                        ans = engine.answer(qid, q_text, context=context, turn_index=idx)
                    except Exception as e:
                        ans = {
                            "question_id": qid,
                            "turn_index": idx,
                            "question": q_text,
                            "status": "error",
                            "clarification": "",
                            "sql": "",
                            "result": [],
                            "analysis": str(e),
                            "charts": [],
                            "context_update": {},
                        }
                    turn_answers.append(ans)
                    if ans.get("context_update"):
                        context.update(ans["context_update"])

                    if ans.get("status") == "error":
                        last_error = ans.get("analysis") or ""
                        last_sql = ans.get("sql") or ""
                        # safe retry: try once more with conservative context hint
                        try:
                            attempts += 1
                            context_retry = dict(context)
                            context_retry["safe_retry"] = True
                            ans2 = engine.answer(qid, q_text, context=context_retry, turn_index=idx)
                            turn_answers.append({"retry": True, **ans2})
                            if ans2.get("context_update"):
                                context.update(ans2["context_update"])
                            if ans2.get("status") != "error":
                                last_error = ""
                                last_sql = ans2.get("sql") or last_sql
                                break
                            else:
                                last_error = ans2.get("analysis") or last_error
                                last_sql = ans2.get("sql") or last_sql
                                failed_flag = True
                        except Exception as e2:
                            last_error = str(e2)
                            failed_flag = True

                final_answer = turn_answers[-1] if turn_answers else {}

                # write per-question JSON with standardized fields
                json_path = os.path.join(result_dir, f"{qid}.json")
                json_obj = {
                    "question_id": qid,
                    "question_type": qtype,
                    "turns": q_turns,
                    "turn_answers": turn_answers,
                    "final_answer": final_answer,
                    "error_type": "",
                    "error_message": "",
                    "attempts": attempts,
                    "last_attempt_ts": datetime.datetime.utcnow().isoformat() + 'Z',
                }
                status = final_answer.get("status", "")
                if status == "error" or failed_flag:
                    if final_answer.get("sql"):
                        etype = "SQL_ERROR"
                    else:
                        etype = "EXEC_ERROR"
                    # prefer explicit analysis text if present
                    emsg = final_answer.get("analysis") or last_error or ""
                    json_obj["error_type"] = etype
                    json_obj["error_message"] = emsg

                    # persist to failed_jobs sqlite
                    try:
                        now = datetime.datetime.utcnow().isoformat() + 'Z'
                        cur_fail.execute(
                            "INSERT OR REPLACE INTO failed_jobs(question_id,question_type,question,status,error_type,error_message,sql,attempts,last_attempt_ts) VALUES(?,?,?,?,?,?,?,?,?)",
                            (qid, qtype, " | ".join(q_turns), status or "error", etype, emsg, final_answer.get("sql") or last_sql, attempts, now),
                        )
                        conn_fail.commit()
                    except Exception as e:
                        logger.exception("Failed to write failed_jobs for %s: %s", qid, e)

                    # append to error log
                    try:
                        logf = os.path.join(result_dir, "batch_errors.log")
                        with open(logf, "a", encoding="utf-8") as lf:
                            lf.write(f"{datetime.datetime.utcnow().isoformat()}Z\t{qid}\t{etype}\t{emsg}\n")
                    except Exception:
                        pass

                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(json_obj, f, ensure_ascii=False, indent=2, default=str)

                out_rows.append(
                    {
                        "编号": qid,
                        "问题类型": qtype,
                        "问题": " | ".join(q_turns),
                        "状态": final_answer.get("status", ""),
                        "澄清问题": final_answer.get("clarification", ""),
                        "执行SQL": final_answer.get("sql", ""),
                        "分析结论": final_answer.get("analysis", ""),
                        "图表路径": "; ".join(final_answer.get("charts", []) or []),
                        "最终结果JSON": json.dumps(final_answer.get("result", []), ensure_ascii=False, default=str),
                        "完整会话JSON": json.dumps(
                            {
                                "turn_answers": turn_answers,
                                "final_answer": final_answer,
                            },
                            ensure_ascii=False,
                            default=str,
                        ),
                    }
                )
    finally:
        db.close()
        try:
            conn_fail.close()
        except Exception:
            pass

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "result_2"
    out_ws.append(
        [
            "编号",
            "问题类型",
            "问题",
            "状态",
            "澄清问题",
            "执行SQL",
            "分析结论",
            "图表路径",
            "最终结果JSON",
            "完整会话JSON",
        ]
    )
    for row in out_rows:
        out_ws.append(
            [
                row["编号"],
                row["问题类型"],
                row["问题"],
                row["状态"],
                row["澄清问题"],
                row["执行SQL"],
                row["分析结论"],
                row["图表路径"],
                row["最终结果JSON"],
                row["完整会话JSON"],
            ]
        )

    # add Failures sheet summarizing failed_jobs
    try:
        conn_fail = sqlite3.connect(failed_db_path)
        cur_fail = conn_fail.cursor()
        cur_fail.execute("SELECT question_id,question_type,error_type,error_message,attempts,last_attempt_ts FROM failed_jobs ORDER BY last_attempt_ts DESC")
        fails = cur_fail.fetchall()
        if fails:
            fail_ws = out_wb.create_sheet(title="Failures")
            fail_ws.append(["编号", "问题类型", "错误类型", "错误信息", "重试次数", "最后尝试时间"])
            for f in fails:
                fail_ws.append(list(f))
        conn_fail.close()
    except Exception:
        pass

    out_wb.save(output_xlsx)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-xlsx", action="append", default=[])
    parser.add_argument(
        "--output-xlsx",
        default=r"D:\BaiduNetdiskDownload\data\smart_query_assistant\result\result_2.xlsx",
    )
    parser.add_argument(
        "--result-dir",
        default=RESULT_DIR,
    )
    args = parser.parse_args()
    input_files = args.input_xlsx or [
        r"D:\BaiduNetdiskDownload\data\附件4：问题汇总.xlsx",
        r"D:\BaiduNetdiskDownload\data\附件6：问题汇总.xlsx",
    ]
    run_batch(input_files, args.output_xlsx, args.result_dir)
    print(f"batch done. output={args.output_xlsx}")


if __name__ == "__main__":
    main()
