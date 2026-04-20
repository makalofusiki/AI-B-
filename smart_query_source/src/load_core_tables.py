from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook
import pymysql

from config import DB_CONFIG


SCHEMA_XLSX = Path(r"D:\BaiduNetdiskDownload\data\附件3：数据库-表名及字段说明.xlsx")
INPUT_DIR_DEFAULT = Path(r"D:\BaiduNetdiskDownload\data\开发用数据表\pre_import_validation\checked")
TEMPLATE_DIR_DEFAULT = Path(r"D:\BaiduNetdiskDownload\data\开发用数据表\pre_import_validation\templates")

SHEET_TABLE_MAP = {
    "核心业绩指标表": "core_performance_indicators_sheet",
    "资产负债表": "balance_sheet",
    "利润表": "income_sheet",
    "现金流量表": "cash_flow_sheet",
}


def load_schema():
    wb = load_workbook(SCHEMA_XLSX, data_only=True)
    out = {}
    for sheet, table in SHEET_TABLE_MAP.items():
        ws = wb[sheet]
        fields = []
        for r in range(2, ws.max_row + 1):
            f = ws.cell(r, 1).value
            if f:
                fields.append(str(f).strip())
        out[table] = fields
    return out


def read_rows(csv_path: Path):
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with csv_path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                return reader.fieldnames or [], rows
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"cannot decode: {csv_path}")


def create_template(template_path: Path, fields):
    template_path.parent.mkdir(parents=True, exist_ok=True)
    with template_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default=str(INPUT_DIR_DEFAULT))
    parser.add_argument("--template-dir", default=str(TEMPLATE_DIR_DEFAULT))
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    template_dir = Path(args.template_dir)
    schema = load_schema()

    conn = pymysql.connect(
        host=DB_CONFIG["host"],
        port=DB_CONFIG["port"],
        user=DB_CONFIG["user"],
        password=DB_CONFIG["password"],
        database=DB_CONFIG["database"],
        charset=DB_CONFIG["charset"],
        autocommit=False,
    )

    try:
        with conn.cursor() as cur:
            for table, fields in schema.items():
                input_csv = input_dir / f"{table}.checked.csv"
                if not input_csv.exists():
                    tpl = template_dir / f"{table}.csv"
                    create_template(tpl, fields)
                    print(f"[SKIP] missing input: {input_csv}")
                    print(f"       template generated: {tpl}")
                    continue

                headers, rows = read_rows(input_csv)
                missing = [c for c in fields if c not in headers]
                if missing:
                    raise RuntimeError(f"{table} missing columns: {missing}")

                cur.execute(f"TRUNCATE TABLE `{table}`")
                if rows:
                    col_sql = ", ".join([f"`{c}`" for c in fields])
                    val_sql = ", ".join(["%s"] * len(fields))
                    sql = f"INSERT INTO `{table}` ({col_sql}) VALUES ({val_sql})"
                    data = []
                    for row in rows:
                        data.append(tuple((row.get(c, "") or None) for c in fields))
                    cur.executemany(sql, data)
                print(f"[OK] {table}: {len(rows)} rows loaded")

        conn.commit()
        print("done.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
